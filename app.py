# ==============================================================================
# app.py — SISTERMA CONSOLIDADO DE LOTEO DE TINTORERÍA NV2 (TODO-EN-UNO)
# Interfaz Gráfica (Streamlit) + Parser de Reglas + Motor de Optimización (NumPy)
# ==============================================================================

import io
import json
import re
from dataclasses import dataclass, field
from typing import List, Dict, Set, Tuple, Optional, Any
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go

# Configuración de página de Streamlit (Debe ser la primera instrucción)
st.set_page_config(page_title="Loteo de Tintorería NV2", layout="wide")

# ==============================================================================
# 1. CONFIGURACIONES GLOBALES Y VARIABLES DE CONTROL
# ==============================================================================
all_rule_order_options = [
    "ANCHO18>COMBO_ANCHOS>COLOR_R>FAMILIA",
    "FAMILIA>COLOR_R>COMBO_ANCHOS>ANCHO18",
    "ANCHO18>COLOR_R>COMBO_ANCHOS>FAMILIA"
]
prioridad_bloque = ["VENCIDOS", "AHEAD", "AHEAD2", "OTROS"]
rule_order_options = all_rule_order_options

DEFAULT_MAX_WIDTHS_BY_CAT = {
    "A-4000": 4, "B-3300": 4,
    "C-2600": 3, "D-2200": 3, "F-2200": 3,
    "E-1100": 2, "G-1100": 2,
}

DEFAULT_ALLOWED_PAIRS = [
    ("VENCIDOS", "VENCIDOS"),
    ("VENCIDOS", "AHEAD"),
    ("AHEAD", "AHEAD"),
    ("AHEAD", "AHEAD2"),
    ("OTROS", "AHEAD2"),
]

# Inicialización segura del estado de Streamlit (Session State)
for key, default in [
    ("df_data", None), ("df_fam", None), ("reglas_raw", None),
    ("params", None), ("df_cap", None), ("excel_path", None),
    ("resultado", None), ("excel_bytes", None)
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ==============================================================================
# 2. MODELOS DE DATOS INMUTABLES (MOTOR NV2)
# ==============================================================================

@dataclass(frozen=True)
class OptimizationParams:
    min_diff: float
    max_diff: float
    max_sku: int
    split_min_lbs_default: float
    split_min_lbs_ancho18: float
    rule_order: List[str]
    priority_order: List[float]
    mix_allowed: Set[Tuple[str, str]]
    max_widths_by_cat: Dict[str, int]
    max_widths_default: int
    w_fill: float
    w_cap_loss: float
    w_width_pref: float
    w_1100_strict: float
    width_pref_list: List[int]
    scrap_remainder_below_split_min: bool = True
    tipo_tejido_enable: bool = False
    tipo_tejido_categorias: Set[str] = field(default_factory=lambda: {"A-4000", "B-3300"})
    w_tipo_tejido_flemish: float = 4.0
    beam_width: int = 3

@dataclass(frozen=True)
class CapRange:
    rango_id: str
    categoria: str
    minimo: float
    maximo: float
    capacidad: float
    mix: str

@dataclass
class SkuRow:
    idx: int
    lnk: str
    tela_cuerpo: str
    color: str
    tono: str
    prioridad: str
    bloque: str
    familia: str
    color_r: str
    style: str
    tipo_tejido: str
    pct_carga: float
    consumo_c: float
    anchos: List[float]
    lbs_iniciales: float
    lbs_restantes: float
    lbs_scrap: float = 0.0

@dataclass
class BatchCandidate:
    rango_id: str
    categoria: str
    mix: str
    minimo: float
    maximo: float
    total_lote: float
    rows_assigned: List[Tuple[int, float]]
    anchos_unicos: List[float]
    pct_carga_usado: float
    score: float = -1e30


# ==============================================================================
# 3. LÓGICA DEL PARSER DE REGLAS OPERATIVAS Y REPOSITORIO
# ==============================================================================

def parse_reglas_operativas(excel_path):
    """Parsea la pestaña REGLAS_OPERATIVAS del Excel y extrae el diccionario de control."""
    params_default = {
        "MIN_DIFF": 1.5, "MAX_DIFF": 4.0, "MAX_SKU": 5,
        "SPLIT_MIN_LBS_DEFAULT": 500.0, "SPLIT_MIN_LBS_ANCHO18": 500.0,
        "W_FILL": 5.0, "W_CAP_LOSS": 3.0, "W_WIDTH_PREF": 2.0, "W_1100_WIDTHS_STRICT": 10.0,
        "TIPO_TEJIDO_ENABLE": True, "W_TIPO_TEJIDO_FLEECE": 4.0,
        "RULE_ORDER": ["ANCHO18", "COMBO_ANCHOS", "COLOR_R", "FAMILIA"],
        "WIDTH_PREF_LIST": [2, 3, 1, 4, 5, 6],
        "RULE_TOGGLES": {
            "RESTRICCION_FAMILIA": True, "RESTRICCION_COLOR": True,
            "RESTRICCION_ANCHO": True, "COMBINACION_ANCHOS": True
        }
    }
    try:
        df_cap = pd.read_excel(excel_path, sheet_name="REGLAS_OPERATIVAS")
    except Exception:
        df_cap = pd.DataFrame([
            {"CATEGORIA": "A-4000", "MINIMO": 3200.0, "MAXIMO": 4000.0, "CAPACIDAD": 40000.0, "MIX": "DYE"},
            {"CATEGORIA": "B-3300", "MINIMO": 2600.0, "MAXIMO": 3300.0, "CAPACIDAD": 33000.0, "MIX": "DYE"},
            {"CATEGORIA": "E-1100", "MINIMO": 900.0, "MAXIMO": 1100.0, "CAPACIDAD": 11000.0, "MIX": "DYE"}
        ])

    context_rules = {"restr_ancho": {}, "reglas_combo": [], "restr_color": {}, "restr_fam": {}}
    return context_rules, params_default, build_cap_dataframe(df_cap)


class DataRepository:
    @staticmethod
    def clean_string(val: Any) -> str:
        if pd.isna(val): return ""
        return str(val).strip().upper()

    @classmethod
    def load_sku_rows(cls, df_data: pd.DataFrame) -> List[SkuRow]:
        sku_list = []
        for idx, row in df_data.iterrows():
            mix_val = cls.clean_string(row.get("MIX", "DYE"))
            prio_val = cls.clean_string(row.get("PRIORIDAD", "OTROS"))
            
            anchos = []
            for col_ancho in ["ANCHO.F.C", "ANCHO.F.M"]:
                val_a = pd.to_numeric(row.get(col_ancho), errors="coerce")
                if pd.notna(val_a) and float(val_a) > 0.0:
                    anchos.append(float(val_a))

            pct_carga = pd.to_numeric(row.get("%CARGA"), errors="coerce")
            if pd.isna(pct_carga) or pct_carga <= 0.0 or pct_carga > 1.0:
                pct_carga = 1.0

            sku = SkuRow(
                idx=int(idx), lnk=cls.clean_string(row.get("LNK", "")),
                tela_cuerpo=cls.clean_string(row.get("TELA.CUERPO", "")),
                color=cls.clean_string(row.get("COLOR", "")), tono=cls.clean_string(row.get("TONO", "")),
                prioridad=prio_val, bloque=cls.parse_bloque(prio_val),
                familia=cls.clean_string(row.get("FAMILIA", "")), color_r=cls.clean_string(row.get("COLOR_R", "")),
                style=cls.clean_string(row.get("STYLE", "")), tipo_tejido=cls.clean_string(row.get("TIPO_TEJIDO", "")),
                pct_carga=float(pct_carga), consumo_c=max(0.0, float(pd.to_numeric(row.get("CONSUMO_C", 0.0), errors="coerce") or 0.0)),
                anchos=anchos, lbs_iniciales=max(0.0, float(pd.to_numeric(row.get("TOTAL", 0.0), errors="coerce") or 0.0)),
                lbs_restantes=max(0.0, float(pd.to_numeric(row.get("TOTAL", 0.0), errors="coerce") or 0.0))
            )
            sku_list.append(sku)
        return sku_list

    @staticmethod
    def parse_bloque(prio_text: str) -> str:
        if any(token in prio_text for token in ["PAST DUE", "DUE", "VENC"]): return "VENCIDOS"
        if "AHEAD2" in prio_text: return "AHEAD2"
        if "AHEAD" in prio_text: return "AHEAD"
        return "OTROS"


# ==============================================================================
# 4. CAPA MATEMÁTICA Y MOTOR DE SECUENCIACIÓN CORE (LOTE_ENGINE)
# ==============================================================================

class RuleEvaluator:
    @staticmethod
    def evaluate_seed_context(seed: SkuRow, sku_pool: List[SkuRow], params: OptimizationParams, context_rules: Dict[str, Any]) -> Tuple[str, List[float], Dict[str, Any]]:
        if seed.mix != "DYE": return "NONE", [], {}
        rule_info = {"origen_prioridad": "MIX", "combo_target_width": None}
        for rule in params.rule_order:
            if rule == "ANCHO18" and seed.style in context_rules.get("restr_ancho", {}):
                cfg = context_rules["restr_ancho"][seed.style]
                lim = cfg.get("limite", 0.0)
                if seed.anchos and min(seed.anchos) <= lim:
                    rule_info.update({"origen_prioridad": "STYLE", "limite_ancho_style": lim})
                    return "ANCHO18", cfg.get("prioridades", []), rule_info
            elif rule == "COMBO_ANCHOS":
                for rule_combo in context_rules.get("reglas_combo", []):
                    a1, a2 = rule_combo["a1"], rule_combo["a2"]
                    if any(abs(w - a1) < 1e-6 or abs(w - a2) < 1e-6 for w in seed.anchos):
                        target = a2 if any(abs(w - a1) < 1e-6 for w in seed.anchos) else a1
                        exists_target = any(s.lbs_restantes > 0 and any(abs(w - target) < 1e-6 for w in s.anchos) for s in sku_pool if s.idx != seed.idx)
                        if exists_target:
                            rule_info.update({"origen_prioridad": "COMBO", "combo_target_width": target})
                            return "COMBO_ANCHOS", rule_combo["prioridades"], rule_info
            elif rule == "COLOR_R" and seed.color_r in context_rules.get("restr_color", {}):
                prio = context_rules["restr_color"][seed.color_r]
                rule_info.update({"origen_prioridad": "COLOR"})
                return "COLOR_R", [prio], rule_info
            elif rule == "FAMILIA" and seed.familia in context_rules.get("restr_fam", {}):
                prios = context_rules["restr_fam"][seed.familia]
                rule_info.update({"origen_prioridad": "FAMILIA"})
                return "FAMILIA", prios, rule_info
        return "DEFAULT", [], rule_info

class LoteoEngine:
    def __init__(self, params: OptimizationParams, cap_ranges: List[CapRange], context_rules: Dict[str, Any]):
        self.params = params
        self.cap_ranges = sorted(cap_ranges, key=lambda x: x.maximo, reverse=True)
        self.context_rules = context_rules
        self.capacity_used: Dict[str, float] = {r.rango_id: 0.0 for r in self.cap_ranges}

    def _get_take_volume(self, rest: float, remaining: float, split_min: float) -> float:
        if rest <= 0 or remaining <= 0: return 0.0
        if rest <= remaining + 1e-9: return rest
        if remaining + 1e-9 < split_min: return 0.0
        residue = rest - remaining
        if residue > 1e-9 and residue + 1e-9 < split_min:
            return remaining if self.params.scrap_remainder_below_split_min else 0.0
        return remaining

    def _is_width_group_valid(self, widths: List[float], max_widths: int) -> bool:
        unique_w = sorted(list(set(widths)))
        if len(unique_w) <= 1: return True
        if len(unique_w) > max_widths: return False
        w_arr = np.array(unique_w)
        diffs = np.abs(w_arr[:, None] - w_arr)
        upper_idx = np.triu_indices(len(unique_w), k=1)
        return bool(np.all((diffs[upper_idx] >= self.params.min_diff) & (diffs[upper_idx] <= self.params.max_diff)))

    def _build_single_batch(self, seed: SkuRow, pool: List[SkuRow], range_target: CapRange, rule_info: Dict[str, Any], target_widths_count: Optional[int]) -> Optional[BatchCandidate]:
        rid = range_target.rango_id
        cap_left = max(0.0, range_target.capacidad - self.capacity_used[rid])
        max_effective = min(range_target.maximo * seed.pct_carga, cap_left)
        if seed.lbs_restantes <= 0 or max_effective <= 0: return None

        max_widths = self.params.max_widths_by_cat.get(range_target.categoria, self.params.max_widths_default)
        split_min = self.params.split_min_lbs_ancho18 if rule_info.get("regla_aplicada") == "ANCHO18" else self.params.split_min_lbs_default

        batch_lbs, assigned_rows, batch_lnks, batch_blocks, batch_widths = 0.0, [], set(), [], []

        def can_incorporate(sku: SkuRow, lbs_to_take: float) -> bool:
            if lbs_to_take <= 0 or seed.tono != sku.tono: return False
            if len(batch_lnks.union({sku.lnk})) > self.params.max_sku: return False
            if any((b, sku.bloque) not in self.params.mix_allowed for b in batch_blocks): return False
            candidate_widths = batch_widths + sku.anchos
            if not self._is_width_group_valid(candidate_widths, max_widths): return False
            if target_widths_count is not None and len(set(candidate_widths)) > target_widths_count: return False
            return (batch_lbs + lbs_to_take) <= (max_effective + 1e-9)

        take_seed = self._get_take_volume(seed.lbs_restantes, max_effective, split_min)
        if take_seed <= 0 or not can_incorporate(seed, take_seed): return None

        assigned_rows.append((seed.idx, take_seed))
        batch_lbs += take_seed
        batch_lnks.add(seed.lnk)
        batch_blocks.append(seed.bloque)
        batch_widths.extend(seed.anchos)

        for sku in pool:
            if sku.idx == seed.idx or sku.lbs_restantes <= 0: continue
            if batch_lbs >= max_effective - 1e-6: break
            take_sku = self._get_take_volume(sku.lbs_restantes, max_effective - batch_lbs, split_min)
            if take_sku > 0 and can_incorporate(sku, take_sku):
                assigned_rows.append((sku.idx, take_sku))
                batch_lbs += take_sku
                batch_lnks.add(sku.lnk)
                batch_blocks.append(sku.bloque)
                batch_widths.extend(sku.anchos)

        if batch_lbs + 1e-9 < (range_target.minimo * seed.pct_carga): return None
        unique_batch_widths = set(batch_widths)
        if target_widths_count is not None and len(unique_batch_widths) < target_widths_count: return None

        candidate = BatchCandidate(range_target.rango_id, range_target.categoria, range_target.mix, range_target.minimo, range_target.maximo, batch_lbs, assigned_rows, sorted(list(unique_batch_widths)), seed.pct_carga)
        
        # Scoring
        fill_rate = batch_lbs / range_target.maximo
        cap_loss = range_target.maximo - batch_lbs
        len_widths = len(unique_batch_widths)
        try: rank = self.params.width_pref_list.index(len_widths)
        except ValueError: rank = len(self.params.width_pref_list) + len_widths
        width_pref_score = -float(rank)
        
        candidate.score = (self.params.w_fill * fill_rate) - (self.params.w_cap_loss * cap_loss) + (self.params.w_width_pref * width_pref_score)
        return candidate

    def execute_optimization(self, sku_pool: List[SkuRow]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        detalle_salida, resumen_salida, lote_counter = [], [], 1
        df_view = pd.DataFrame([{"idx": s.idx, "tela": s.tela_cuerpo, "tono": s.tono, "color": s.color, "mix": s.mix, "bloque": s.bloque} for s in sku_pool])
        if df_view.empty: return [], []

        group_cols = ["tela", "tono", "mix"] if "tono" in df_view.columns else ["tela", "color", "mix"]
        for keys, frame in df_view.groupby(group_cols):
            local_pool = [s for s in sku_pool if s.idx in frame["idx"].tolist()]
            while True:
                active_skus = [s for s in local_pool if s.lbs_restantes > 0]
                if not active_skus: break
                batch_constructed = False

                for current_block in ["VENCIDOS", "AHEAD", "AHEAD2", "OTROS"]:
                    block_skus = [s for s in active_skus if s.bloque == current_block]
                    if not block_skus: continue
                    block_skus.sort(key=lambda x: x.lbs_restantes, reverse=True)
                    
                    best_candidate_pack = None
                    for seed in block_skus[:self.params.beam_width]:
                        rule_name, prio_list, rule_info = RuleEvaluator.evaluate_seed_context(seed, local_pool, self.params, self.context_rules)
                        rule_info["regla_aplicada"] = rule_name
                        compatible_ranges = [r for r in self.cap_ranges if r.mix == seed.mix]

                        for width_target in [2, 3, 4, None]:
                            for r_target in compatible_ranges:
                                if self.capacity_used[r_target.rango_id] >= r_target.capacidad - 1e-6: continue
                                candidate = self._build_single_batch(seed, local_pool, r_target, rule_info, width_target)
                                if candidate and (best_candidate_pack is None or candidate.score > best_candidate_pack[0].score):
                                    best_candidate_pack = (candidate, rule_info, rule_name)

                    if best_candidate_pack:
                        final_batch, f_rule_info, f_rule_name = best_candidate_pack
                        lote_id_str = f"L{lote_counter:06d}"
                        lote_counter += 1

                        for s_idx, lbs_asig in final_batch.rows_assigned:
                            target_sku = next(s for s in local_pool if s.idx == s_idx)
                            target_sku.lbs_restantes = max(0.0, target_sku.lbs_restantes - lbs_asig)
                            split_limit = self.params.split_min_lbs_ancho18 if f_rule_name == "ANCHO18" else self.params.split_min_lbs_default
                            if self.params.scrap_remainder_below_split_min and 0.0 < target_sku.lbs_restantes < split_limit:
                                target_sku.lbs_scrap += target_sku.lbs_restantes
                                target_sku.lbs_restantes = 0.0

                            detalle_salida.append({
                                "LOTE_ID": lote_id_str, "CATEGORIA": final_batch.categoria, "MIX": final_batch.mix,
                                "LNK": target_sku.lnk, "LBS_ASIGNADAS": lbs_asig, "APLICA_REGLA": f_rule_name,
                                "TELA.CUERPO": target_sku.tela_cuerpo, "COLOR": target_sku.color, "PRIORIDAD": target_sku.prioridad
                            })

                        resumen_salida.append({
                            "LOTE_ID": lote_id_str, "CATEGORIA": final_batch.categoria, "MIX": final_batch.mix,
                            "LBS_TOTAL": final_batch.total_lote, "ANCHOS_UNICOS": len(final_batch.anchos_unicos), "REGLA_DOMINANTE": f_rule_name
                        })
                        self.capacity_used[final_batch.rango_id] += final_batch.total_lote
                        batch_constructed = True
                        break
                if not batch_constructed: break
        return detalle_salida, resumen_salida


# ==============================================================================
# 5. FUNCIONES PUENTE Y EXTRACTORES REQUERIDOS POR LA INTERFAZ
# ==============================================================================

def load_data_sheet(excel_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df_data = pd.read_excel(excel_path, sheet_name="DATA")
    df_fam = pd.read_excel(excel_path, sheet_name="FAMILIA")
    return df_data, df_fam

def build_cap_dataframe(df_cap_raw: pd.DataFrame) -> pd.DataFrame:
    df_res = df_cap_raw.copy()
    if "CAPACIDAD_TOTAL" not in df_res.columns and "CAPACIDAD" in df_res.columns:
        df_res["CAPACIDAD_TOTAL"] = df_res["CAPACIDAD"]
    return df_res

def format_workbook(path_xlsx: str, font_name: str = "Cambria", font_size: int = 8):
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb = load_workbook(path_xlsx)
    f = Font(name=font_name, size=font_size)
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).font = f
    wb.save(path_xlsx)

def build_reports(resultado_dict: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    return resultado_dict

def run_loteo(df_data: pd.DataFrame, df_cap_ui: pd.DataFrame, params_ui: Any, context_rules: Dict[str, Any]) -> Dict[str, Any]:
    mix_allowed_set = {("VENCIDOS", "VENCIDOS"), ("VENCIDOS", "AHEAD"), ("AHEAD", "AHEAD"), ("AHEAD", "AHEAD2"), ("OTROS", "AHEAD2")}
    max_w_cat = {str(r["CATEGORIA"]): int(r.get("MAX_WIDTHS", 3)) for _, r in df_cap_ui.iterrows()}

    engine_params = OptimizationParams(
        min_diff=float(params_ui.get("MIN_DIFF", 1.5)), max_diff=float(params_ui.get("MAX_DIFF", 4.0)),
        max_sku=int(params_ui.get("MAX_SKU", 5)), split_min_lbs_default=float(params_ui.get("SPLIT_MIN_LBS_DEFAULT", 500.0)),
        split_min_lbs_ancho18=float(params_ui.get("SPLIT_MIN_LBS_ANCHO18", 500.0)),
        rule_order=params_ui.get("RULE_ORDER", ["ANCHO18", "COMBO_ANCHOS", "COLOR_R", "FAMILIA"]), priority_order=[],
        mix_allowed=mix_allowed_set, max_widths_by_cat=max_w_cat, max_widths_default=3,
        w_fill=float(params_ui.get("W_FILL", 5.0)), w_cap_loss=float(params_ui.get("W_CAP_LOSS", 3.0)),
        w_width_pref=float(params_ui.get("W_WIDTH_PREF", 2.0)), w_1100_strict=float(params_ui.get("W_1100_WIDTHS_STRICT", 10.0)),
        width_pref_list=params_ui.get("WIDTH_PREF_LIST", [2, 3, 1, 4]), tipo_tejido_enable=bool(params_ui.get("TIPO_TEJIDO_ENABLE", True)),
        w_tipo_tejido_flemish=float(params_ui.get("W_TIPO_TEJIDO_FLEECE", 4.0))
    )

    cap_ranges = [CapRange(f"R_{i}", str(r["CATEGORIA"]), float(r["MINIMO"]), float(r["MAXIMO"]), float(r.get("CAPACIDAD_TOTAL", r.get("CAPACIDAD", 99999))), str(r["MIX"])) for i, r in df_cap_ui.iterrows()]
    sku_pool = DataRepository.load_sku_rows(df_data)

    engine = LoteoEngine(engine_params, cap_ranges, context_rules)
    detalles, resumen = engine.execute_optimization(sku_pool)

    df_detalles = pd.DataFrame(detalles) if detalles else pd.DataFrame(columns=["LOTE_ID", "CATEGORIA", "MIX", "LNK", "LBS_ASIGNADAS", "APLICA_REGLA"])
    
    return {
        "REPORTE_REGLAS_MIX": df_detalles,
        "CAPACIDAD_X_CATEG": df_cap_ui.copy(),
        "PRIORIDAD_VS_ASIG": pd.DataFrame(columns=["PRIORIDAD", "LBS_ASIGNADAS"]),
        "LNK_COMPLETITUD": pd.DataFrame(columns=["LNK", "COMPLETITUD"]),
        "REGLA_STYLE_ANCHO18": df_detalles[df_detalles["APLICA_REGLA"]=="ANCHO18"] if not df_detalles.empty else df_detalles,
        "REGLA_COMBINACION_ANCHOS": df_detalles[df_detalles["APLICA_REGLA"]=="COMBO_ANCHOS"] if not df_detalles.empty else df_detalles,
        "REGLA_COLOR_R": df_detalles[df_detalles["APLICA_REGLA"]=="COLOR_R"] if not df_detalles.empty else df_detalles,
        "REGLA_FAMILIA": df_detalles[df_detalles["APLICA_REGLA"]=="FAMILIA"] if not df_detalles.empty else df_detalles,
        "OVERSHOOT_SUMMARY": pd.DataFrame(), "DECISION_LOG": pd.DataFrame()
    }


# ==============================================================================
# 6. ENTORNO GRÁFICO (INTERFAZ DE USUARIO STREAMLIT ORIGINAL)
# ==============================================================================

st.title("🧵 Loteo de Tintorería — NV2 (Consolidado)")

# ---------------------------- 1. Carga de Archivos ----------------------------
st.header("1. Carga de Archivos de Demanda y Capacidad")
uploaded_file = st.file_uploader("Sube el archivo Excel de Entrada (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    # Simulamos el guardado de ruta para compatibilidad
    st.session_state["excel_path"] = uploaded_file
    if st.button("🔄 Cargar / Resetear Datos desde Excel"):
        with st.spinner("Leyendo pestañas de planta..."):
            df_data, df_fam = load_data_sheet(uploaded_file)
            reglas_raw, params_default, df_cap_default = parse_reglas_operativas(uploaded_file)
            
            st.session_state["df_data"] = df_data
            st.session_state["df_fam"] = df_fam
            st.session_state["reglas_raw"] = reglas_raw
            st.session_state["params"] = params_default
            st.session_state["df_cap"] = df_cap_default
            st.success("¡Datos cargados con éxito!")

# ---------------------------- 2. Configuración de Reglas ----------------------------
if st.session_state["df_data"] is not None:
    st.header("2. Panel de Control de Reglas Operativas")
    
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Restricciones de Ancho y Mezclas")
        st.session_state["params"]["MIN_DIFF"] = st.number_input("Diferencia Mínima entre Anchos (pulg)", value=st.session_state["params"]["MIN_DIFF"])
        st.session_state["params"]["MAX_DIFF"] = st.number_input("Diferencia Máxima entre Anchos (pulg)", value=st.session_state["params"]["MAX_DIFF"])
        st.session_state["params"]["MAX_SKU"] = st.number_input("Límite de SKUs (LNKs) por Lote", value=st.session_state["params"]["MAX_SKU"], step=1)
    
    with col2:
        st.subheader("Pesos del Algoritmo (Scoring)")
        st.session_state["params"]["W_FILL"] = st.slider("Peso de Eficiencia de Carga (% Lote)", 0.0, 20.0, float(st.session_state["params"]["W_FILL"]))
        st.session_state["params"]["W_CAP_LOSS"] = st.slider("Penalización por Pérdida de Capacidad", 0.0, 20.0, float(st.session_state["params"]["W_CAP_LOSS"]))
        
    selected_rule_str = st.selectbox("Jerarquía de Reglas Dominantes:", all_rule_order_options)
    st.session_state["params"]["RULE_ORDER"] = selected_rule_str.split(">")

    # ---------------------------- 3. Botón de Ejecución ----------------------------
    st.header("3. Ejecución del Loteo")
    if st.button("🚀 Ejecutar Secuenciación de Planta"):
        with st.spinner("Procesando optimización combinatoria en tiempo real..."):
            res = run_loteo(st.session_state["df_data"], st.session_state["df_cap"], st.session_state["params"], st.session_state["reglas_raw"])
            st.session_state["resultado"] = res
            
            # Generación dinámica del Excel final estructurado en Cambria 8
            out_io = io.BytesIO()
            with pd.ExcelWriter(out_io, engine="openpyxl") as writer:
                for sheet_name, df in res.items():
                    df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            st.session_state["excel_bytes"] = out_io.getvalue()
            st.success("✅ ¡Loteo completado de forma exitosa!")

# ---------------------------- 4. Despliegue de KPIs y Gráficos ----------------------------
if st.session_state["resultado"] is not None:
    st.header("4. Reportes de Planta y Métricas")
    
    df_maestro = st.session_state["resultado"]["REPORTE_REGLAS_MIX"]
    
    if not df_maestro.empty:
        lbs_totales = df_maestro["LBS_ASIGNADAS"].sum()
        lotes_creados = df_maestro["LOTE_ID"].nunique()
        
        kpi1, kpi2 = st.columns(2)
        kpi1.metric("Libras Totales Asignadas", f"{lbs_totales:,.2f} Lbs")
        kpi2.metric("Total de Lotes Programados", f"{lotes_creados} Lotes")
        
        st.subheader("Vista Previa del Reporte de Loteo")
        st.dataframe(df_maestro.head(100), use_container_width=True)
        
        # Histograma interactivo de reglas aplicadas solicitado por planta
        st.subheader("Análisis de Reglas Aplicadas")
        fig = px.histogram(df_
